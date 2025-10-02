# =========================
# PARTE 1 — BASE + PERSISTENCIA + DESPIECE COMPLETO
# =========================
import streamlit as st
import pandas as pd
import numpy as np
import math
import os
from io import BytesIO

# -------- PERSISTENCIA DE STOCK --------
DEFAULT_STOCK = "stock.csv"  # semilla del repo (no se pisa)
APP_DIR = os.path.dirname(os.path.abspath(__file__))
RUNTIME_STOCK = os.path.join(APP_DIR, "stock_runtime.csv")  # archivo persistente local
ARCHIVO_STOCK = RUNTIME_STOCK  # se usa siempre este handler

# -------- Google Sheets (opcional, recomendado) --------
GSHEET_TAB = "Stock"  # nombre de pestaña dentro del doc
_GS_ENABLED = False
_STOCK_SHEET_ID = None
try:
    # Activamos GS solo si existen ambos secrets
    if hasattr(st, "secrets") and ("gcp_service_account" in st.secrets) and ("STOCK_SHEET_ID" in st.secrets):
        from google.oauth2 import service_account
        import gspread

        _GCP_CREDS = service_account.Credentials.from_service_account_info(
            st.secrets["gcp_service_account"],
            scopes=[
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive",
            ],
        )
        _STOCK_SHEET_ID = st.secrets["STOCK_SHEET_ID"]
        _GS_ENABLED = True
except Exception:
    _GS_ENABLED = False  # si falla algo, quedamos en modo local sin mostrar nada

def _ensure_local_seed():
    """Inicializa el CSV local persistente desde DEFAULT_STOCK si no existe."""
    if not os.path.exists(RUNTIME_STOCK):
        try:
            seed_df = pd.read_csv(DEFAULT_STOCK)
        except Exception:
            seed_df = pd.DataFrame(columns=["Categoría", "Tipo", "Unidad", "Stock"])
        seed_df.to_csv(RUNTIME_STOCK, index=False)

def _ensure_gsheet_seed():
    """Inicializa la pestaña 'Stock' si está vacía o no existe, copiando desde DEFAULT_STOCK."""
    try:
        gc = gspread.authorize(_GCP_CREDS)
        sh = gc.open_by_key(_STOCK_SHEET_ID)
        try:
            ws = sh.worksheet(GSHEET_TAB)
        except gspread.WorksheetNotFound:
            ws = sh.add_worksheet(title=GSHEET_TAB, rows=1000, cols=10)

        values = ws.get_all_values()
        if not values:
            try:
                seed_df = pd.read_csv(DEFAULT_STOCK)
            except Exception:
                seed_df = pd.DataFrame(columns=["Categoría", "Tipo", "Unidad", "Stock"])
            ws.clear()
            ws.update([seed_df.columns.tolist()] + seed_df.values.tolist())
        else:
            header = values[0]
            expected = ["Categoría", "Tipo", "Unidad", "Stock"]
            if header != expected:
                df = pd.DataFrame(values[1:], columns=header)
                df.columns = [c.strip() for c in df.columns]
                for col in expected:
                    if col not in df.columns:
                        df[col] = "" if col != "Stock" else 0
                df = df[expected]
                ws.clear()
                ws.update([expected] + df.values.tolist())
    except Exception:
        pass  # si falla Sheets, se usa local

# Inicialización de la semilla según el modo
if _GS_ENABLED:
    _ensure_gsheet_seed()
else:
    _ensure_local_seed()

# ======= LECTURA/ESCRITURA PERSISTENTE =======
def _gs_read_stock() -> pd.DataFrame:
    gc = gspread.authorize(_GCP_CREDS)
    sh = gc.open_by_key(_STOCK_SHEET_ID)
    ws = sh.worksheet(GSHEET_TAB)
    values = ws.get_all_values()
    if not values:
        return pd.DataFrame(columns=["Categoría", "Tipo", "Unidad", "Stock"])
    header = values[0]
    data = values[1:]
    df = pd.DataFrame(data, columns=header)
    if "Stock" in df.columns:
        df["Stock"] = pd.to_numeric(df["Stock"], errors="coerce").fillna(0.0)
    for col in ["Categoría", "Tipo", "Unidad"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    return df

def _gs_write_stock(df: pd.DataFrame):
    gc = gspread.authorize(_GCP_CREDS)
    sh = gc.open_by_key(_STOCK_SHEET_ID)
    try:
        ws = sh.worksheet(GSHEET_TAB)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=GSHEET_TAB, rows=1000, cols=10)
    df2 = df.copy()
    for col in ["Categoría", "Tipo", "Unidad"]:
        if col in df2.columns:
            df2[col] = df2[col].astype(str).str.strip()
    if "Stock" in df2.columns:
        df2["Stock"] = pd.to_numeric(df2["Stock"], errors="coerce").fillna(0.0)
    ws.clear()
    ws.update([df2.columns.tolist()] + df2.values.tolist())

def cargar_stock():
    if _GS_ENABLED:
        try:
            return _gs_read_stock()
        except Exception:
            pass  # si falla, caemos a local silenciosamente
    # Local fallback
    df = pd.read_csv(RUNTIME_STOCK)
    for col in ["Categoría", "Tipo", "Unidad"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    if "Stock" in df.columns:
        df["Stock"] = pd.to_numeric(df["Stock"], errors="coerce").fillna(0.0)
    return df

def guardar_stock(df: pd.DataFrame):
    # Intentar Sheets si está habilitado
    if _GS_ENABLED:
        try:
            _gs_write_stock(df)
            return
        except Exception:
            pass
    # Local fallback
    df2 = df.copy()
    if "Stock" in df2.columns:
        df2["Stock"] = pd.to_numeric(df2["Stock"], errors="coerce").fillna(0.0)
    df2.to_csv(RUNTIME_STOCK, index=False)

# --- Fracción de goma espuma por modelo ---
def calcular_fraccion_goma_espuma(modelo):
    fracciones = {
        "Silla Mora": 1/15,
        "Silla Franca": 1/15,
        "Silla Xanaes": 1/15,
        "Silla Gala": 1/12,
        "Silla Eva": 1/12,
        "Silla Luna": 1/20
    }
    return fracciones.get(modelo, 0)

# --- Tela por área (rollo base 140 x 50 cm) ---
TELA_LARGO_BASE_CM = 140
TELA_ANCHO_BASE_CM = 50
TELA_AREA_BASE_CM2 = TELA_LARGO_BASE_CM * TELA_ANCHO_BASE_CM  # 7000
TELA_TIPO_BASE_PREFERIDO = "Rollo 140x50"

ANCHO_TELA_CM = {
    ("Silla Franca", "Tela asiento"): 46,
    ("Silla Luna", "Tela asiento"): 46,
    ("Silla Mora", "Tela asiento"): 46,
    ("Silla Mora", "Tela respaldo"): 23,
    ("Silla Xanaes", "Tela asiento"): 46,
    ("Silla Xanaes", "Tela respaldo"): 40,
    ("Silla Gala", "Tela asiento"): 55,
    ("Silla Gala", "Tela respaldo"): 69,
    ("Banquetas Umma", "Tela asiento"): 46,
    ("Banquetas Umma", "Tela respaldo"): 39,
}

def localizar_tipo_tela(df_stock, keywords, preferido=TELA_TIPO_BASE_PREFERIDO):
    tipos = df_stock[df_stock["Categoría"] == "Tela"]["Tipo"].astype(str)
    for t in tipos:
        s = t.lower().replace(" ", "")
        for kw in keywords:
            if kw in s:
                return t
    if ((df_stock["Categoría"] == "Tela") & (df_stock["Tipo"] == preferido)).any():
        return preferido
    return preferido

def es_respaldo(tipo_str: str) -> bool:
    return "respaldo" in str(tipo_str).lower()

def calcular_barras_usadas(cortes):
    total_cm = sum(cortes)
    return math.ceil(total_cm / 600)

# --- DESPIECE COMPLETO ---
df_despiece = pd.DataFrame([
    # Silla Franca
    ["Silla Franca", "Caño", '1 1/2"', 42, 2],
    ["Silla Franca", "Caño", '1 1/2"', 85, 2],
    ["Silla Franca", "Caño", '15x25mm', 50, 2],
    ["Silla Franca", "Caño", '1/2"', 33.5, 2],
    ["Silla Franca", "Caño", '1/2"', 37, 1],
    ["Silla Franca", "Caño", '10x20mm', 35, 3],
    ["Silla Franca", "Tela", 'Tela asiento', 55, 1],   # 55x46
    ["Silla Franca", "Tornillo", '4½ x 45mm', 0, 2],
    ["Silla Franca", "Tornillo", '4½ x 40mm', 0, 2],

    # Silla Luna
    ["Silla Luna", "Caño", '1"', 105, 2],
    ["Silla Luna", "Caño", '1"', 66, 2],
    ["Silla Luna", "Caño", '3/4"', 37, 2],
    ["Silla Luna", "Caño", 'Hierro del 6 liso', 34, 4],
    ["Silla Luna", "Tela", 'Tela asiento', 40, 1],     # 40x46
    ["Silla Luna", "Tornillo", '4½ x 35mm', 0, 3],

    # Silla Mora
    ["Silla Mora", "Caño", '1 1/2"', 42, 2],
    ["Silla Mora", "Caño", '1 1/2"', 85, 2],
    ["Silla Mora", "Caño", '15x25mm', 50, 2],
    ["Silla Mora", "Caño", '1/2"', 33.5, 2],
    ["Silla Mora", "Caño", '1/2"', 37, 1],
    ["Silla Mora", "Chapa", 'Diagonal 2 perforaciones', 0, 2],
    ["Silla Mora", "Tela", 'Tela asiento', 55, 1],      # 55x46
    ["Silla Mora", "Tela", 'Tela respaldo', 39, 1],     # 39x23
    ["Silla Mora", "Madera", 'Contratapa', 25, 1],      # 25x39
    ["Silla Mora", "Tornillo", '4½ x 45mm', 0, 2],
    ["Silla Mora", "Tornillo", '4½ x 40mm', 0, 2],
    ["Silla Mora", "Tornillo", '4½ x 16mm', 0, 2],
    ["Silla Mora", "Tornillo", '4½ x 20mm', 0, 4],

    # Silla Xanaes
    ["Silla Xanaes", "Caño", '1 1/2"', 42, 2],
    ["Silla Xanaes", "Caño", '1 1/2"', 85, 2],
    ["Silla Xanaes", "Caño", '15x25mm', 50, 2],
    ["Silla Xanaes", "Caño", '1/2"', 33.5, 2],
    ["Silla Xanaes", "Caño", '1/2"', 37, 1],
    ["Silla Xanaes", "Chapa", 'Rectangular 3 perforaciones', 0, 2],
    ["Silla Xanaes", "Tela", 'Tela asiento', 55, 1],    # 55x46
    ["Silla Xanaes", "Tela", 'Tela respaldo', 40, 1],   # 40x40
    ["Silla Xanaes", "Madera", 'Contratapa', 25, 1],    # 25x39
    ["Silla Xanaes", "Tornillo", '4½ x 45mm', 0, 2],
    ["Silla Xanaes", "Tornillo", '4½ x 40mm', 0, 2],
    ["Silla Xanaes", "Tornillo", '4½ x 16mm', 0, 2],

    # Silla Eva (sin tela)
    ["Silla Eva", "Caño", '1"', 43, 2],
    ["Silla Eva", "Caño", '1"', 87, 2],
    ["Silla Eva", "Caño", '15x25mm', 34.5, 3],
    ["Silla Eva", "Caño", '15x25mm', 35.5, 1],
    ["Silla Eva", "Caño", '1/2"', 35.5, 1],
    ["Silla Eva", "Caño", '1/2"', 35.5, 3],
    ["Silla Eva", "Tornillo", '4½ x 45mm', 0, 2],
    ["Silla Eva", "Tornillo", '4½ x 35mm', 0, 1],

    # Silla Gala
    ["Silla Gala", "Caño", '1"', 71, 2],
    ["Silla Gala", "Caño", '1"', 100, 2],
    ["Silla Gala", "Caño", '15x25mm', 35, 1],
    ["Silla Gala", "Caño", '15x25mm', 31, 1],
    ["Silla Gala", "Caño", '1/2"', 31.5, 1],
    ["Silla Gala", "Caño", '1/2"', 31, 1],
    ["Silla Gala", "Tela", 'Tela asiento', 53, 1],      # 53x55
    ["Silla Gala", "Tela", 'Tela respaldo', 78, 1],     # 78x69
    ["Silla Gala", "Tornillo", '4½ x 35mm', 0, 4],

    # Banquetas Umma
    ["Banquetas Umma", "Caño", '1"', 114, 2],
    ["Banquetas Umma", "Caño", '1"', 70, 2],
    ["Banquetas Umma", "Caño", '3/4"', 35.5, 1],
    ["Banquetas Umma", "Caño", '3/4"', 29.5, 1],
    ["Banquetas Umma", "Caño", '15x25mm', 42, 2],
    ["Banquetas Umma", "Caño", '1/2"', 45.5, 2],
    ["Banquetas Umma", "Tela", 'Tela asiento', 50, 1],   # 50x46
    ["Banquetas Umma", "Tela", 'Tela respaldo', 25, 1],  # 25x39
    ["Banquetas Umma", "Madera", 'Contratapa', 25, 1],   # 25x39
    ["Banquetas Umma", "Tornillo", '4½ x 45mm', 0, 3],
    ["Banquetas Umma", "Tornillo", '4½ x 16mm', 0, 2],
    ["Banquetas Umma", "Tornillo", '4½ x 20mm', 0, 4],

    # Mesa
    ["Mesa", "Caño", '2"', 74.5, 4],
    ["Mesa", "Caño", '20x20', 60, 4],
    ["Mesa", "Caño", '20x40', 4, 4],
    ["Mesa", "Caño", '20x20', 1100, 4],
    ["Mesa", "Madera", 'Contratapa', 0, 1],

    # Perchero
    ["Perchero", "Caño", '3/4"', 40, 4],
    ["Perchero", "Caño", '3/4"', 50, 4],
    ["Perchero", "Caño", '3/4"', 50, 1],
    ["Perchero", "Caño", '1 1/4"', 130, 1],
], columns=["Modelo", "Categoría", "Tipo", "Largo (cm)", "Cantidad"])
# =========================
# PARTE 2 — UI + SIMULACIÓN + GUARDADO
# =========================
st.set_page_config(page_title="Sistema de Despiece", layout="wide")
st.title("📐 Sistema de Despiece y Stock – DELARTE")

menu = st.sidebar.radio("Menú", ["📐 Despiece", "📦 Stock"])

if menu == "📐 Despiece":
    modelos = df_despiece["Modelo"].unique()
    modelo_seleccionado = st.selectbox("Seleccioná un modelo", modelos)
    cantidad = st.number_input("Cantidad a producir", min_value=1, step=1)
    cliente = st.text_input("Cliente")
    color_tela = st.text_input("Color de tapizado")
    color_cano = st.text_input("Color de caño")

    df_modelo = df_despiece[df_despiece["Modelo"] == modelo_seleccionado].copy()
    df_modelo["Total"] = df_modelo["Cantidad"] * cantidad

    # --- GOMA ESPUMA (visual: 1 por silla / stock: fraccionado) ---
    fraccion_goma = calcular_fraccion_goma_espuma(modelo_seleccionado)
    if fraccion_goma > 0:
        df_modelo = pd.concat([
            df_modelo,
            pd.DataFrame([{
                "Modelo": modelo_seleccionado,
                "Categoría": "Goma espuma",
                "Tipo": "Plancha",
                "Largo (cm)": 0,
                "Cantidad": 1,
                "Total": cantidad
            }])
        ], ignore_index=True)

    # --- Simulación de stock ---
    df_stock = cargar_stock()

    # Detectar renglones base de tela
    tipo_tela_asiento = localizar_tipo_tela(df_stock, keywords=["asiento"])
    tipo_tela_respaldo = localizar_tipo_tela(df_stock, keywords=["respaldo"])

    # Diccionario de stock
    stock_dict = {(fila["Categoría"], fila["Tipo"]): fila["Stock"] for _, fila in df_stock.iterrows()}

    # Ancho visible (numérico o NaN)
    def ancho_visible_row(row):
        if row["Categoría"] == "Tela":
            return float(ANCHO_TELA_CM.get((row["Modelo"], row["Tipo"]), np.nan))
        if row["Categoría"] == "Madera" and str(row["Tipo"]).strip().lower() == "contratapa":
            return 39.0  # 25x39
        return np.nan

    df_modelo["Ancho (cm)"] = pd.to_numeric(df_modelo.apply(ancho_visible_row, axis=1), errors="coerce")

    # Stock actual por fila
    def obtener_stock_actual(row):
        if row["Categoría"] == "Tela":
            if es_respaldo(row["Tipo"]):
                return stock_dict.get(("Tela", tipo_tela_respaldo), 0.0)
            else:
                return stock_dict.get(("Tela", tipo_tela_asiento), 0.0)
        return stock_dict.get((row["Categoría"], row["Tipo"]), 0.0)

    df_modelo["Stock Actual"] = df_modelo.apply(obtener_stock_actual, axis=1)

    # --- Agrupación de consumo ---
    consumo_real = {}
    tela_area_asiento_cm2 = 0.0
    tela_area_respaldo_cm2 = 0.0

    for _, row in df_modelo.iterrows():
        clave = (row["Categoría"], row["Tipo"])
        if row["Categoría"] == "Caño":
            largo_total = row["Total"] * row["Largo (cm)"]
            consumo_real[clave] = consumo_real.get(clave, 0.0) + float(largo_total)  # cm lineales
        elif row["Categoría"] == "Goma espuma":
            consumo_real[clave] = consumo_real.get(clave, 0.0) + float(row["Total"] * fraccion_goma)
        elif row["Categoría"] == "Tela":
            # área = largo x ancho
            ancho_val = row["Ancho (cm)"]
            if pd.isna(ancho_val):
                ancho_val = float(ANCHO_TELA_CM.get((row["Modelo"], row["Tipo"]), 0.0))
            area = float(row["Largo (cm)"]) * float(ancho_val)  # cm²
            if es_respaldo(row["Tipo"]):
                tela_area_respaldo_cm2 += area * row["Total"]
            else:
                tela_area_asiento_cm2 += area * row["Total"]
        else:
            # Madera (Contratapa) y otros: unidades
            consumo_real[clave] = consumo_real.get(clave, 0.0) + float(row["Total"])

    # Convertir áreas a fracciones de rollo
    fraccion_tela_asiento = tela_area_asiento_cm2 / TELA_AREA_BASE_CM2 if tela_area_asiento_cm2 > 0 else 0.0
    fraccion_tela_respaldo = tela_area_respaldo_cm2 / TELA_AREA_BASE_CM2 if tela_area_respaldo_cm2 > 0 else 0.0

    if fraccion_tela_asiento > 0:
        consumo_real[("Tela", tipo_tela_asiento)] = consumo_real.get(("Tela", tipo_tela_asiento), 0.0) + fraccion_tela_asiento
    if fraccion_tela_respaldo > 0:
        consumo_real[("Tela", tipo_tela_respaldo)] = consumo_real.get(("Tela", tipo_tela_respaldo), 0.0) + fraccion_tela_respaldo

    # --- Estados ---
    estados = []
    for _, row in df_modelo.iterrows():
        if row["Categoría"] == "Caño":
            clave = (row["Categoría"], row["Tipo"])
            stock_actual = float(stock_dict.get(clave, 0.0))
            caños_necesarios = float(consumo_real.get(clave, 0.0)) / 600.0  # 600 cm = 1 barra
            estado = "✅ OK" if stock_actual >= caños_necesarios else "❌ Faltante"
        elif row["Categoría"] == "Tela":
            if es_respaldo(row["Tipo"]):
                stock_actual = float(stock_dict.get(("Tela", tipo_tela_respaldo), 0.0))
                estado = "✅ OK" if stock_actual >= fraccion_tela_respaldo else "❌ Faltante"
            else:
                stock_actual = float(stock_dict.get(("Tela", tipo_tela_asiento), 0.0))
                estado = "✅ OK" if stock_actual >= fraccion_tela_asiento else "❌ Faltante"
        else:
            clave = (row["Categoría"], row["Tipo"])
            stock_actual = float(stock_dict.get(clave, 0.0))
            estado = "✅ OK" if stock_actual >= float(consumo_real.get(clave, 0.0)) else "❌ Faltante"
        estados.append(estado)

    df_modelo["Estado"] = estados

    # --- Visualización ---
    st.subheader("🧩 Despiece y Simulación de Stock")
    columnas_vista = ["Categoría", "Tipo", "Largo (cm)", "Ancho (cm)", "Cantidad", "Total", "Stock Actual", "Estado"]
    df_vista = df_modelo[columnas_vista].copy()
    st.dataframe(df_vista, use_container_width=True)

    # --- Guardado de stock (PERSISTENTE) ---
    if st.button("💾 Guardar Producción y Actualizar Stock", key="guardar_produccion"):
        for clave, consumo in consumo_real.items():
            if clave[0] == "Caño":
                caños_usados = float(consumo) / 600.0
                mask = (df_stock["Categoría"] == clave[0]) & (df_stock["Tipo"] == clave[1])
                if mask.any():
                    df_stock.loc[mask, "Stock"] = pd.to_numeric(df_stock.loc[mask, "Stock"], errors="coerce").fillna(0.0) - caños_usados
            else:
                mask = (df_stock["Categoría"] == clave[0]) & (df_stock["Tipo"] == clave[1])
                if mask.any():
                    df_stock.loc[mask, "Stock"] = pd.to_numeric(df_stock.loc[mask, "Stock"], errors="coerce").fillna(0.0) - float(consumo)

        guardar_stock(df_stock)  # guarda en Sheets si hay, o en stock_runtime.csv
        st.success("✅ Producción guardada y stock actualizado.")

elif menu == "📦 Stock":
    st.subheader("📦 Stock de Materiales")
    df_stock = cargar_stock()
    edited_df = st.data_editor(df_stock, num_rows="dynamic", use_container_width=True)
    if st.button("💾 Guardar Cambios en Stock", key="guardar_stock"):
        guardar_stock(edited_df)
        st.success("✅ Cambios guardados correctamente.")
# =========================
# PARTE 3 — EXPORTACIÓN A EXCEL + DESCARGA
# =========================
def exportar_excel(df, modelo, cantidad, cliente, color_tela, color_cano):
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, Alignment
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Despiece"

    # Configuración de impresión (A4 horizontal, 1 página)
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.3, bottom=0.3)
    ws.print_options.horizontalCentered = True

    # Encabezado
    encabezado = [
        ["Cliente:", cliente],
        ["Modelo:", modelo],
        ["Cantidad:", cantidad],
        ["Color de tapizado:", color_tela],
        ["Color de caño:", color_cano]
    ]
    for i, fila in enumerate(encabezado, start=1):
        ws.cell(row=i, column=1, value=fila[0]).font = Font(bold=True)
        ws.cell(row=i, column=2, value=fila[1])

    # Tabla de despiece
    columnas = ["Categoría", "Tipo", "Largo (cm)", "Ancho (cm)", "Cantidad", "Total", "Stock Actual", "Estado"]
    datos = df[columnas].copy()

    start_row = len(encabezado) + 2
    for r_idx, row in enumerate(dataframe_to_rows(datos, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            val = "" if (value is None or (isinstance(value, float) and math.isnan(value))) else value
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if r_idx == start_row:
                cell.font = Font(bold=True)

    # Ancho personalizado de columnas
    columnas_personalizadas = {1: 18, 2: 22, 3: 14, 4: 14, 5: 10, 6: 18, 7: 16, 8: 14}
    for col, width in columnas_personalizadas.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

# --- BOTÓN PARA IMPRIMIR DESPIECE ---
try:
    if 'menu' in locals() and menu == "📐 Despiece" and 'df_modelo' in locals() and not df_modelo.empty:
        excel_bytes = exportar_excel(df_modelo, modelo_seleccionado, cantidad, cliente, color_tela, color_cano)
        st.download_button(
            label="🖨️ Imprimir Despiece",
            data=excel_bytes,
            file_name=f"Despiece_{modelo_seleccionado}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="descargar_excel"
        )
except Exception:
    pass
